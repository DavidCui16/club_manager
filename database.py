import sqlite3
import os
import sys
from datetime import datetime

if getattr(sys, "frozen", False):
    base_dir = os.path.dirname(sys.executable)
else:
    base_dir = os.path.dirname(__file__)

DB_PATH = os.path.join(base_dir, "club.db")


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            student_id TEXT UNIQUE,
            phone TEXT,
            email TEXT,
            department TEXT,
            join_date TEXT NOT NULL DEFAULT (date('now')),
            status TEXT NOT NULL DEFAULT 'active',
            notes TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            event_date TEXT NOT NULL,
            location TEXT,
            description TEXT,
            max_participants INTEGER DEFAULT 0,
            fee REAL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'upcoming'
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS event_participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL,
            member_id INTEGER NOT NULL,
            FOREIGN KEY (event_id) REFERENCES events(id) ON DELETE CASCADE,
            FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE CASCADE,
            UNIQUE(event_id, member_id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS finances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('income', 'expense')),
            amount REAL NOT NULL,
            category TEXT NOT NULL,
            description TEXT,
            date TEXT NOT NULL DEFAULT (date('now')),
            member_id INTEGER,
            event_id INTEGER,
            FOREIGN KEY (member_id) REFERENCES members(id) ON DELETE SET NULL,
            FOREIGN KEY (event_id) REFERENCES events(id) ON DELETE SET NULL
        )
    """)

    conn.commit()
    conn.close()


# ---------- Members ----------
def add_member(name, student_id="", phone="", email="", department="", notes=""):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO members (name, student_id, phone, email, department, notes) VALUES (?,?,?,?,?,?)",
        (name, student_id, phone, email, department, notes),
    )
    conn.commit()
    conn.close()


def update_member(member_id, name, student_id, phone, email, department, status, notes):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE members SET name=?, student_id=?, phone=?, email=?, department=?, status=?, notes=? WHERE id=?",
        (name, student_id, phone, email, department, status, notes, member_id),
    )
    conn.commit()
    conn.close()


def delete_member(member_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id=?", (member_id,))
    conn.commit()
    conn.close()


def get_all_members():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return rows


def search_members(keyword):
    conn = get_conn()
    kw = f"%{keyword}%"
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM members WHERE name LIKE ? OR student_id LIKE ? OR department LIKE ? OR phone LIKE ? ORDER BY id DESC",
        (kw, kw, kw, kw),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


# ---------- Events ----------
def add_event(name, event_date, location="", description="", max_participants=0, fee=0):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO events (name, event_date, location, description, max_participants, fee) VALUES (?,?,?,?,?,?)",
        (name, event_date, location, description, max_participants, fee),
    )
    conn.commit()
    conn.close()


def update_event(event_id, name, event_date, location, description, max_participants, fee, status):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE events SET name=?, event_date=?, location=?, description=?, max_participants=?, fee=?, status=? WHERE id=?",
        (name, event_date, location, description, max_participants, fee, status, event_id),
    )
    conn.commit()
    conn.close()


def delete_event(event_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM events WHERE id=?", (event_id,))
    conn.commit()
    conn.close()


def get_all_events():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """SELECT e.*, (SELECT COUNT(*) FROM event_participants WHERE event_id = e.id) AS participant_count
           FROM events e ORDER BY e.event_date DESC"""
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def search_events(keyword):
    conn = get_conn()
    kw = f"%{keyword}%"
    cur = conn.cursor()
    cur.execute(
        """SELECT e.*, (SELECT COUNT(*) FROM event_participants WHERE event_id = e.id) AS participant_count
           FROM events e WHERE e.name LIKE ? OR e.location LIKE ? ORDER BY e.event_date DESC""",
        (kw, kw),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def add_participant(event_id, member_id):
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO event_participants (event_id, member_id) VALUES (?,?)",
            (event_id, member_id),
        )
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def remove_participant(event_id, member_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "DELETE FROM event_participants WHERE event_id=? AND member_id=?",
        (event_id, member_id),
    )
    conn.commit()
    conn.close()


def get_event_participants(event_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """SELECT m.* FROM members m
           JOIN event_participants ep ON m.id = ep.member_id
           WHERE ep.event_id = ? ORDER BY m.name""",
        (event_id,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


# ---------- Finances ----------
def add_finance(fin_type, amount, category, description="", member_id=None, event_id=None):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO finances (type, amount, category, description, member_id, event_id) VALUES (?,?,?,?,?,?)",
        (fin_type, amount, category, description, member_id, event_id),
    )
    conn.commit()
    conn.close()


def update_finance(fin_id, fin_type, amount, category, description):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE finances SET type=?, amount=?, category=?, description=? WHERE id=?",
        (fin_type, amount, category, description, fin_id),
    )
    conn.commit()
    conn.close()


def delete_finance(fin_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM finances WHERE id=?", (fin_id,))
    conn.commit()
    conn.close()


def get_all_finances():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """SELECT f.*, m.name AS member_name, e.name AS event_name
           FROM finances f
           LEFT JOIN members m ON f.member_id = m.id
           LEFT JOIN events e ON f.event_id = e.id
           ORDER BY f.date DESC, f.id DESC"""
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def get_finance_summary():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT SUM(amount) FROM finances WHERE type='income'")
    total_income = cur.fetchone()[0] or 0
    cur.execute("SELECT SUM(amount) FROM finances WHERE type='expense'")
    total_expense = cur.fetchone()[0] or 0
    conn.close()
    return total_income, total_expense, total_income - total_expense


def get_member_dict():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM members WHERE status='active' ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return {row["name"]: row["id"] for row in rows}


def get_event_dict():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, name FROM events ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return {row["name"]: row["id"] for row in rows}


def bulk_add_members(members_data):
    conn = get_conn()
    cur = conn.cursor()
    count = 0
    skipped = 0
    for m in members_data:
        name = m.get("name", "").strip()
        if not name:
            continue
        try:
            cur.execute(
                "INSERT INTO members (name, student_id, phone, email, department, notes) VALUES (?,?,?,?,?,?)",
                (name, m.get("student_id", ""), m.get("phone", ""),
                 m.get("email", ""), m.get("department", ""), m.get("notes", "")),
            )
            count += 1
        except sqlite3.IntegrityError:
            skipped += 1
    conn.commit()
    conn.close()
    return count, skipped


def import_members_from_excel(filepath):
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return 0, 0

    headers = [str(h).strip() if h else "" for h in rows[0]]

    column_map = {
        "姓名": "name", "name": "name",
        "学号": "student_id", "student_id": "student_id",
        "电话": "phone", "phone": "phone",
        "邮箱": "email", "email": "email",
        "院系": "department", "department": "department",
        "备注": "notes", "notes": "notes",
    }

    col_indices = {}
    for i, h in enumerate(headers):
        key = column_map.get(h)
        if key and key not in col_indices:
            col_indices[key] = i

    if "name" not in col_indices:
        wb.close()
        raise ValueError("Excel 文件缺少「姓名」列，请检查表头是否为: 姓名, 学号, 电话, 邮箱, 院系, 备注")

    members = []
    for row in rows[1:]:
        member = {}
        for field, idx in col_indices.items():
            val = row[idx] if idx < len(row) else ""
            member[field] = str(val).strip() if val is not None else ""
        if member.get("name"):
            members.append(member)

    wb.close()
    return bulk_add_members(members)
